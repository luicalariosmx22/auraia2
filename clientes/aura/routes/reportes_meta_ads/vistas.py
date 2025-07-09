from . import reportes_meta_ads_bp
from flask import render_template, request, redirect, url_for, flash, jsonify
from clientes.aura.utils.supabase_client import supabase
from datetime import datetime, timedelta
import pandas as pd
import io
import base64

@reportes_meta_ads_bp.route('/reportes', methods=['GET', 'POST'])
def vista_reportes_meta_ads(nombre_nora=None):
    if nombre_nora is None:
        nombre_nora = request.args.get('nombre_nora')
    # Obtener cuentas que coincidan con nombre_nora o nombre_visible
    cuentas_ads = supabase.table('meta_ads_cuentas').select('*').eq('nombre_nora', nombre_nora).execute().data or []
    reportes = supabase.table('meta_ads_reportes').select('*').eq('nombre_nora', nombre_nora).order('fecha_envio', desc=True).limit(30).execute().data or []
    if request.method == 'POST' and 'archivo_reporte' in request.files:
        archivo = request.files['archivo_reporte']
        if archivo.filename:
            flash('Reporte subido correctamente (lógica pendiente de implementar)', 'success')
            return redirect(url_for('reportes_meta_ads.vista_reportes_meta_ads', nombre_nora=nombre_nora))
    return render_template('reportes_meta_ads/reportes_meta_ads.html', nombre_nora=nombre_nora, cuentas_ads=cuentas_ads, reportes=reportes)

@reportes_meta_ads_bp.route('/generar_especifico', methods=['POST'])
def generar_reporte_especifico(nombre_nora=None):
    """Genera un reporte específico para una cuenta publicitaria"""
    try:
        if nombre_nora is None:
            nombre_nora = request.view_args.get('nombre_nora')
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No se enviaron datos'}), 400
        
        cuenta_id = data.get('cuenta_id')
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        nombre_cliente = data.get('nombre_cliente')
        plataforma = data.get('plataforma')
        
        # Validar datos requeridos
        if not all([cuenta_id, fecha_desde, fecha_hasta, nombre_cliente]):
            return jsonify({'success': False, 'error': 'Faltan datos requeridos'}), 400
        
        # Convertir fechas
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'error': 'Formato de fecha inválido'}), 400
        
        # Validar rango de fechas
        if fecha_desde_obj > fecha_hasta_obj:
            return jsonify({'success': False, 'error': 'La fecha desde no puede ser mayor a la fecha hasta'}), 400
        
        # Obtener datos de la cuenta
        cuenta_info = supabase.table('meta_ads_cuentas').select('*').eq('id_cuenta_publicitaria', cuenta_id).eq('nombre_visible', nombre_nora).execute().data
        
        if not cuenta_info:
            return jsonify({'success': False, 'error': f'No se encontró la cuenta {cuenta_id}'}), 404
        
        cuenta = cuenta_info[0]
        
        # Obtener datos de anuncios para el rango de fechas
        anuncios_data = supabase.table('meta_ads_anuncios_detalle').select('*').eq('id_cuenta_publicitaria', cuenta_id).gte('fecha_inicio', fecha_desde).lte('fecha_fin', fecha_hasta).execute().data
        
        if not anuncios_data:
            return jsonify({'success': False, 'error': f'No se encontraron datos para el período {fecha_desde} - {fecha_hasta}'}), 404
        
        # Generar reporte Excel
        reporte_excel = generar_excel_reporte(anuncios_data, cuenta, fecha_desde, fecha_hasta)
        
        # Guardar reporte en la base de datos
        reporte_id = guardar_reporte_bd(
            nombre_nora=nombre_nora,
            cuenta_id=cuenta_id,
            nombre_cliente=nombre_cliente,
            plataforma=plataforma,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            archivo_excel=reporte_excel
        )
        
        # Generar URL de descarga
        download_url = url_for('reportes_meta_ads.descargar_reporte', reporte_id=reporte_id, nombre_nora=nombre_nora)
        
        return jsonify({
            'success': True,
            'message': f'Reporte generado exitosamente para {nombre_cliente}',
            'download_url': download_url,
            'reporte_id': reporte_id,
            'total_anuncios': len(anuncios_data),
            'periodo': f'{fecha_desde} - {fecha_hasta}'
        })
        
    except Exception as e:
        print(f"❌ Error generando reporte específico: {e}")
        return jsonify({'success': False, 'error': f'Error interno del servidor: {str(e)}'}), 500

def generar_excel_reporte(anuncios_data, cuenta, fecha_desde, fecha_hasta):
    """Genera un archivo Excel con los datos del reporte"""
    try:
        # Crear DataFrame
        df = pd.DataFrame(anuncios_data)
        
        # Configurar columnas relevantes
        columnas_reporte = [
            'nombre_campana', 'nombre_conjunto_anuncios', 'nombre_anuncio',
            'fecha_inicio', 'fecha_fin', 'impresiones', 'clics', 'gasto',
            'ctr', 'cpm', 'cpc', 'conversiones', 'costo_por_conversion'
        ]
        
        # Filtrar columnas que existen
        columnas_existentes = [col for col in columnas_reporte if col in df.columns]
        df_reporte = df[columnas_existentes]
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja principal con datos
            df_reporte.to_excel(writer, sheet_name='Reporte Anuncios', index=False)
            
            # Hoja de resumen
            resumen_data = {
                'Cliente': [cuenta.get('nombre_cliente', 'N/A')],
                'Plataforma': [cuenta.get('tipo_plataforma', 'N/A')],
                'ID Cuenta': [cuenta.get('id_cuenta_publicitaria', 'N/A')],
                'Período': [f'{fecha_desde} - {fecha_hasta}'],
                'Total Anuncios': [len(anuncios_data)],
                'Gasto Total': [df['gasto'].sum() if 'gasto' in df.columns else 0],
                'Impresiones Totales': [df['impresiones'].sum() if 'impresiones' in df.columns else 0],
                'Clics Totales': [df['clics'].sum() if 'clics' in df.columns else 0],
                'Fecha Generación': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            }
            
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        print(f"❌ Error generando Excel: {e}")
        raise

def guardar_reporte_bd(nombre_nora, cuenta_id, nombre_cliente, plataforma, fecha_desde, fecha_hasta, archivo_excel):
    """Guarda el reporte en la base de datos"""
    try:
        # Codificar archivo en base64
        archivo_base64 = base64.b64encode(archivo_excel).decode('utf-8')
        
        # Insertar en la tabla de reportes
        reporte_data = {
            'nombre_visible': nombre_nora,
            'id_cuenta_publicitaria': cuenta_id,
            'nombre_cliente': nombre_cliente,
            'tipo_plataforma': plataforma,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'archivo_excel': archivo_base64,
            'fecha_generacion': datetime.now().isoformat(),
            'estado': 'completado',
            'tipo_reporte': 'especifico'
        }
        
        result = supabase.table('meta_ads_reportes').insert(reporte_data).execute()
        
        if result.data:
            return result.data[0]['id']
        else:
            raise Exception("No se pudo insertar el reporte")
            
    except Exception as e:
        print(f"❌ Error guardando reporte en BD: {e}")
        raise

@reportes_meta_ads_bp.route('/descargar/<int:reporte_id>')
def descargar_reporte(reporte_id, nombre_nora=None):
    """Descarga un reporte específico"""
    try:
        if nombre_nora is None:
            nombre_nora = request.view_args.get('nombre_nora')
        
        # Obtener reporte de la base de datos
        reporte = supabase.table('meta_ads_reportes').select('*').eq('id', reporte_id).eq('nombre_visible', nombre_nora).execute().data
        
        if not reporte:
            flash('Reporte no encontrado', 'error')
            return redirect(url_for('reportes_meta_ads.vista_reportes_meta_ads', nombre_nora=nombre_nora))
        
        reporte_data = reporte[0]
        
        # Decodificar archivo Excel
        archivo_excel = base64.b64decode(reporte_data['archivo_excel'])
        
        # Generar nombre del archivo
        nombre_archivo = f"reporte_{reporte_data['nombre_cliente']}_{reporte_data['fecha_desde']}_{reporte_data['fecha_hasta']}.xlsx"
        
        # Enviar archivo
        from flask import send_file
        return send_file(
            io.BytesIO(archivo_excel),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        print(f"❌ Error descargando reporte: {e}")
        flash(f'Error al descargar reporte: {str(e)}', 'error')
        return redirect(url_for('reportes_meta_ads.vista_reportes_meta_ads', nombre_nora=nombre_nora))

@reportes_meta_ads_bp.route('/data_empresa/<cuenta_id>')
def data_empresa(cuenta_id, nombre_nora=None):
    """Obtiene los datos de una empresa específica"""
    try:
        if nombre_nora is None:
            nombre_nora = request.view_args.get('nombre_nora')
        
        # Obtener datos de la cuenta
        cuenta_info = supabase.table('meta_ads_cuentas').select('*').eq('id_cuenta_publicitaria', cuenta_id).eq('nombre_visible', nombre_nora).execute().data
        
        if not cuenta_info:
            return jsonify({'error': 'No se encontró la empresa'}), 404
        
        cuenta = cuenta_info[0]
        
        return jsonify({
            'nombre_cliente': cuenta.get('nombre_cliente', 'N/A'),
            'tipo_plataforma': cuenta.get('tipo_plataforma', 'N/A'),
            'id_cuenta_publicitaria': cuenta.get('id_cuenta_publicitaria', 'N/A'),
            'account_status': cuenta.get('account_status', 'N/A'),
            'fecha_creacion': cuenta.get('fecha_creacion', 'N/A'),
            'moneda': cuenta.get('moneda', 'N/A')
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo datos de empresa: {e}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500

@reportes_meta_ads_bp.route('/generar_especifico', methods=['POST'])
def generar_especifico(nombre_nora=None):
    """Genera un reporte específico para una cuenta publicitaria"""
    try:
        if nombre_nora is None:
            nombre_nora = request.view_args.get('nombre_nora')
        
        # Obtener datos del request
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400
        
        cuenta_id = data.get('cuenta_id')
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        nombre_cliente = data.get('nombre_cliente', '')
        plataforma = data.get('plataforma', '')
        
        # Validar datos requeridos
        if not all([cuenta_id, fecha_desde, fecha_hasta]):
            return jsonify({'error': 'Faltan datos requeridos: cuenta_id, fecha_desde, fecha_hasta'}), 400
        
        # Validar formato de fechas
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
        
        # Validar rango de fechas
        if fecha_desde_obj > fecha_hasta_obj:
            return jsonify({'error': 'La fecha desde no puede ser mayor a la fecha hasta'}), 400
        
        # Validar que el rango no sea mayor a 1 año
        if (fecha_hasta_obj - fecha_desde_obj).days > 365:
            return jsonify({'error': 'El rango de fechas no puede ser mayor a 1 año'}), 400
        
        # Verificar que la cuenta existe y pertenece al usuario
        cuenta_info = supabase.table('meta_ads_cuentas').select('*').eq('id_cuenta_publicitaria', cuenta_id).eq('nombre_visible', nombre_nora).execute().data
        
        if not cuenta_info:
            return jsonify({'error': 'Cuenta publicitaria no encontrada o sin acceso'}), 404
        
        cuenta = cuenta_info[0]
        
        # Consultar datos de Meta Ads para la cuenta específica
        meta_ads_data = supabase.table('meta_ads_data') \
            .select('*') \
            .eq('account_id', cuenta_id) \
            .gte('date_start', fecha_desde) \
            .lte('date_stop', fecha_hasta) \
            .execute().data
        
        if not meta_ads_data:
            return jsonify({'error': 'No se encontraron datos para el rango de fechas especificado'}), 404
        
        # Generar archivo Excel
        import io
        import base64
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte {nombre_cliente}"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Encabezados
        headers = [
            'Fecha', 'Campaña', 'Conjunto de anuncios', 'Anuncio',
            'Impresiones', 'Clics', 'CTR (%)', 'CPC (USD)',
            'Gasto (USD)', 'Conversiones', 'CPA (USD)', 'ROAS'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        row = 2
        for data_row in meta_ads_data:
            ws.cell(row=row, column=1, value=data_row.get('date_start', ''))
            ws.cell(row=row, column=2, value=data_row.get('campaign_name', ''))
            ws.cell(row=row, column=3, value=data_row.get('adset_name', ''))
            ws.cell(row=row, column=4, value=data_row.get('ad_name', ''))
            ws.cell(row=row, column=5, value=int(data_row.get('impressions', 0)))
            ws.cell(row=row, column=6, value=int(data_row.get('clicks', 0)))
            ws.cell(row=row, column=7, value=round(float(data_row.get('ctr', 0)), 2))
            ws.cell(row=row, column=8, value=round(float(data_row.get('cpc', 0)), 2))
            ws.cell(row=row, column=9, value=round(float(data_row.get('spend', 0)), 2))
            ws.cell(row=row, column=10, value=int(data_row.get('actions', 0)))
            ws.cell(row=row, column=11, value=round(float(data_row.get('cost_per_action', 0)), 2))
            ws.cell(row=row, column=12, value=round(float(data_row.get('roas', 0)), 2))
            row += 1
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        archivo_excel_base64 = base64.b64encode(output.getvalue()).decode('utf-8')
        
        # Guardar reporte en base de datos
        reporte_data = {
            'nombre_visible': nombre_nora,
            'id_cuenta_publicitaria': cuenta_id,
            'nombre_cliente': nombre_cliente,
            'tipo_plataforma': plataforma,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'fecha_generacion': datetime.now().isoformat(),
            'archivo_excel': archivo_excel_base64,
            'total_registros': len(meta_ads_data),
            'estado': 'completado'
        }
        
        resultado = supabase.table('meta_ads_reportes').insert(reporte_data).execute()
        
        if resultado.data:
            reporte_id = resultado.data[0]['id']
            download_url = url_for('reportes_meta_ads.descargar', reporte_id=reporte_id, nombre_nora=nombre_nora)
            
            return jsonify({
                'success': True,
                'message': f'Reporte generado exitosamente para {nombre_cliente}',
                'reporte_id': reporte_id,
                'download_url': download_url,
                'total_registros': len(meta_ads_data),
                'periodo': f'{fecha_desde} a {fecha_hasta}'
            })
        else:
            return jsonify({'error': 'Error al guardar el reporte en la base de datos'}), 500
            
    except Exception as e:
        print(f"❌ Error generando reporte específico: {e}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
